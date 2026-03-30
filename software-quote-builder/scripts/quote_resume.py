#!/usr/bin/env python3
"""Resume helpers for historical software quote workspaces."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from quote_workspace_common import parse_project_dir_name

TEMPLATE_HEADERS = [
    '序号',
    '一级模块',
    '二级模块',
    '功能点',
    '功能说明',
    '预估工时（人天）',
    '单价（元/人天）',
    '小计（元）',
    '备注',
]
QUOTE_COLUMNS = ['预估工时（人天）', '单价（元/人天）', '小计（元）']
SUMMARY_MARKERS = {'合计', '总计'}


def _trim(values: list[Any]) -> list[Any]:
    trimmed = list(values)
    while trimmed and (trimmed[-1] is None or str(trimmed[-1]).strip() == ''):
        trimmed.pop()
    return trimmed


def _cell_text(value: Any) -> str:
    return '' if value is None else str(value).strip()


def _as_float(value: Any, default: float = 0.0) -> float:
    if value in (None, ''):
        return default
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(',', '')
    return float(text) if text else default


def discover_base_output_xlsx(base_project_dir: Path) -> Path | None:
    output_dir = base_project_dir / 'output'
    if not output_dir.is_dir():
        return None
    candidates = sorted(
        (path for path in output_dir.glob('*.xlsx') if path.is_file()),
        key=lambda path: (path.stat().st_mtime, path.name),
        reverse=True,
    )
    return candidates[0] if candidates else None


def detect_project_name(base_project_dir: Path, title: str) -> str:
    title = (title or '').strip()
    suffix = '功能清单报价表'
    if title.endswith(suffix):
        name = title[: -len(suffix)].strip()
        if name:
            return name
    stem, _timestamp = parse_project_dir_name(base_project_dir.name)
    return stem


def recover_quote_payload_from_xlsx(xlsx_path: Path, fallback_project_dir: Path | None = None) -> dict[str, object]:
    workbook = load_workbook(xlsx_path, data_only=True)
    sheet = workbook['报价表'] if '报价表' in workbook.sheetnames else workbook[workbook.sheetnames[0]]
    rows = [_trim([cell for cell in row]) for row in sheet.iter_rows(values_only=True)]
    rows = [row for row in rows if row]
    if len(rows) < 2:
        raise ValueError(f'workbook {xlsx_path} missing quote table rows')

    title = _cell_text(rows[0][0]) if rows[0] else ''
    headers = [_cell_text(value) for value in rows[1]]
    data_rows = []
    special_notes_enabled = False

    for row in rows[2:]:
        first = _cell_text(row[0]) if row else ''
        if first in SUMMARY_MARKERS:
            continue
        if first == '特殊说明':
            special_notes_enabled = True
            break
        data_rows.append(row)

    project_dir = fallback_project_dir or xlsx_path.parent.parent
    project_name = detect_project_name(project_dir, title)
    rate = 800.0

    if headers[: len(TEMPLATE_HEADERS)] == TEMPLATE_HEADERS:
        items = []
        for idx, row in enumerate(data_rows, start=1):
            if not any(_cell_text(cell) for cell in row):
                continue
            row = row + [None] * (len(TEMPLATE_HEADERS) - len(row))
            rate = _as_float(row[6], rate)
            item = {
                'seq': int(_as_float(row[0], idx)),
                'module_l1': _cell_text(row[1]),
                'module_l2': _cell_text(row[2]),
                'feature': _cell_text(row[3]),
                'description': _cell_text(row[4]),
                'estimated_days': _as_float(row[5], 0.0),
            }
            note = _cell_text(row[8])
            if note:
                item['note'] = note
            items.append(item)
        return {
            'project_name': project_name,
            'mode': 'template',
            'day_rate': rate,
            'special_notes_enabled': special_notes_enabled,
            'items': items,
        }

    if headers[-3:] != QUOTE_COLUMNS:
        raise ValueError(f'workbook {xlsx_path} does not contain a resumable quote table')

    base_columns = headers[:-3]
    items = []
    base_rows = []
    for idx, row in enumerate(data_rows, start=1):
        row = row + [None] * (len(headers) - len(row))
        base_row = [row[col] for col in range(len(base_columns))]
        base_rows.append(base_row)
        rate = _as_float(row[len(base_columns) + 1], rate)
        item = {
            'seq': idx,
            'estimated_days': _as_float(row[len(base_columns)], 0.0),
        }
        items.append(item)

    return {
        'project_name': project_name,
        'mode': 'reuse',
        'day_rate': rate,
        'special_notes_enabled': special_notes_enabled,
        'base_columns': base_columns,
        'base_rows': base_rows,
        'items': items,
    }


def build_source_regenerated_stub(project_name: str, source_paths: list[Path]) -> dict[str, object]:
    return {
        'project_name': project_name,
        'mode': 'template',
        'special_notes_enabled': True,
        'source_files': [path.name for path in source_paths],
        'assumptions': ['需要基于 source/ 原始材料重新整理清单后再继续报价调整'],
        'items': [],
    }
